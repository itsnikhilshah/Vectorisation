from pinecone import Pinecone, ServerlessSpec, PodSpec
from sentence_transformers import SentenceTransformer
import torch
import spacy
from datetime import datetime
from time import time
import pickle
import configparser
import os
import pandas as pd
import glob
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from openpyxl.styles import PatternFill
import re

def color_identical_cells_in_excel(folder_path):
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for headers
    blue_fill= PatternFill(start_color='3297A8', end_color='3297A8', fill_type='solid')

    # Iterate over all .xlsx files in the folder
    for file in glob.glob(f"{folder_path}/*.xlsx"):
        workbook = openpyxl.load_workbook(file)
        
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]

            # Apply header color to the first row
            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    cell.fill = header_fill

            # Collect columns indices that start with "Course"
            course_columns = []
            if worksheet.max_row > 0:
                for col in worksheet.iter_cols(1, worksheet.max_column):
                    if col[0].value and re.fullmatch(r"Course \d+", col[0].value):
                        course_columns.append(col[0].column)

            course_name_columns=[]

            if worksheet.max_row > 0:
                for col in worksheet.iter_cols(1, worksheet.max_column):
                    if col[0].value and re.fullmatch(r"Course Name \d+", col[0].value):
                        course_name_columns.append(col[0].column)




            # Iterate through each cell in identified Course columns
            for col_index in course_columns:
                cells = [cell for cell in worksheet.iter_cols(min_col=col_index, max_col=col_index, min_row=2, max_row=worksheet.max_row)]
                for cell_group in cells:
                    for cell in cell_group:
                        for compare_group in cells:
                            for compare_cell in compare_group:
                                if cell != compare_cell and cell.value == compare_cell.value:
                                    cell.fill = light_green_fill
                                    compare_cell.fill = light_green_fill
            
            for col_index in course_name_columns:
                cells = [cell for cell in worksheet.iter_cols(min_col=col_index, max_col=col_index, min_row=2, max_row=worksheet.max_row)]
                for cell_group in cells:
                    for cell in cell_group:
                        for compare_group in cells:
                            for compare_cell in compare_group:
                                if cell != compare_cell and cell.value == compare_cell.value:
                                    cell.fill = blue_fill
                                    compare_cell.fill = blue_fill
        
        # Save the changes
        workbook.save(file)

def pivot_table(folder_path, output_filename="pivot_table_output.xlsx"):
    # Check if the specified folder exists
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' does not exist.")
        return
    
    # Gather all Excel files in the folder
    files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    if not files:
        print("No Excel files found in the folder.")
        return

    # List to collect data frames for all courses
    course_details = []

    # Process each file
    for file in files:
        # Read the Excel file
        df = pd.read_excel(file)

        # Identify columns for 'Course Name' and 'Course Link'
        course_name_cols = [col for col in df.columns if 'Course Name' in col]
        course_link_cols = [col.replace('Name', 'Link') for col in course_name_cols if col.replace('Name', 'Link') in df.columns]

        # Collect data from identified columns
        for name_col, link_col in zip(course_name_cols, course_link_cols):
            # Ensure both columns exist before accessing them
            if name_col in df.columns and link_col in df.columns:
                course_details.extend(df[[name_col, link_col]].dropna().values.tolist())

    # Create a DataFrame from the list of course details
    if course_details:
        course_details_df = pd.DataFrame(course_details, columns=['Course Name', 'Course Link'])

        # Group data by 'Course Name' and find the most common 'Course Link' for each 'Course Name'
        course_pivot = course_details_df.groupby('Course Name')['Course Link'].agg(
            lambda x: x.mode().iloc[0] if not x.mode().empty else pd.NA
        ).reset_index()

        # Count frequencies of course names
        course_pivot['Frequency'] = course_details_df['Course Name'].value_counts().reindex(course_pivot['Course Name']).values

        # Sort the DataFrame by 'Frequency' in descending order
        course_pivot = course_pivot.sort_values(by='Frequency', ascending=False)

        # Define the output file path within the same folder
        output_file = os.path.join(folder_path, output_filename)

        # Save the pivot table to an Excel file
        try:
            course_pivot.to_excel(output_file, sheet_name='Pivot Table', index=False)
            print(f"Pivot table saved to {output_file}")
        except Exception as e:
            print(f"Failed to save the pivot table: {e}")
    else:
        print("No course data available to create a pivot table.")

    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    workbook = openpyxl.load_workbook(output_file)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                cell.fill = header_fill
    workbook.save(output_file)

def clean_course_desc(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    corrected_lines = []
    for line in lines:
        if line.strip() and (line[0].isalnum() and "-" in line):
            corrected_lines.append(line)
        elif corrected_lines: 
            print('Correction made')
            corrected_lines[-1] = corrected_lines[-1].strip() + " " + line

    with open(file_path, 'w', encoding='utf-8') as file:
        file.writelines(corrected_lines)


class PineconeInteraction:
    def __init__(self):
        config=configparser.ConfigParser()
        config.read('settings.ini')
        self.secret_key=config['DEFAULT']['SECRET_KEY']
        index_name = 'serc-index'
        pc=Pinecone(api_key=self.secret_key)

        existing_indexes = pc.list_indexes().names()
        if index_name in existing_indexes:
            print(f"Index '{index_name}' already exists.")
        else:
            pc.create_index(
                name=index_name,
                dimension=1024,
                metric="cosine",
                spec=PodSpec(
                environment="gcp-starter"
                )
            )
        
        self.index=pc.Index(index_name)

        device = 'cuda' if torch.cuda.is_available() else 'cpu'
        model_name = 'sentence-transformers/all-roberta-large-v1'
        self.model = SentenceTransformer(model_name)
        self.corpus_sentences = []
    
    def corpus_embeddings(self,filename):
        clean_course_desc(filename)
        txt_file = open(filename,"r", encoding='utf8') 
        txt_lines=txt_file.readlines() #reading each line individually

        course_id = []
        course_descriptions = []

        for data in txt_lines: #splitting the course id and course description
            split_data = data.split(" -> ")
            if len(split_data) >= 2:
                course_id.append(split_data[0])
                course_descriptions.append(split_data[1])
            else:
                print(f"Warning: '->' not found in item: {data}")
        
        self.corpus_sentences=course_descriptions
        #calculating corpus embeddings
        corpus_embedding = self.model.encode(self.corpus_sentences).tolist()

        #chunk the data into batches of 200
        chunk_size = 200

        #storing the calculated values into pickle file to reduce need to constantly process
        with open('data\\course_data.pkl','wb') as f:
            pickle.dump((course_id,course_descriptions),f)


        for i in range(0, len(corpus_embedding), chunk_size):
            data_chunk = list(zip(course_id[i:i+chunk_size], corpus_embedding[i:i+chunk_size])) #preparing the data chunk
            
            #uploading to vector datastore
            self.index.upsert(data_chunk,namespace="course_descriptions")


            print("Chunk "+str(i)+" has been upserted")


        
    def compare_KSAB(self):
        folderpath = "data/roles"
        print('start')

        if not os.path.exists("data\\combined_KSAB_excel"):
            os.mkdir('data\\combined_KSAB_excel')

        with open('data\\course_id_to_name.pkl', 'rb') as f:
            course_id_to_name = pickle.load(f)

        for filename in glob.glob(os.path.join(folderpath, '*.xlsx')):
            df_roles = pd.read_excel(filename)
            query_sentences = df_roles['description : String'].tolist()
            query_embeddings = self.model.encode(query_sentences)

            similar_pairs = []

            # Calculate cosine similarity for each pair of embeddings
            for i, embedding1 in enumerate(query_embeddings):
                for j in range(i + 1, len(query_embeddings)):
                    similarity = cosine_similarity([embedding1], [query_embeddings[j]])[0][0]
                    if similarity > 0.75:
                        similar_pairs.append((i, j))

            # Sort pairs for efficient processing
            similar_pairs.sort()

            # Combine the sentences in pairs and modify the query_sentences list
            replaced_indices = set()
            for i, j in similar_pairs:
                if i not in replaced_indices and j not in replaced_indices:
                    concatenated_sentence = query_sentences[i] + " " + query_sentences[j]
                    query_sentences[i] = concatenated_sentence
                    replaced_indices.add(j)

            # Remove the replaced indices from the query_sentences
            query_sentences = [sentence for idx, sentence in enumerate(query_sentences) if idx not in replaced_indices]
            query_embeddings = self.model.encode(query_sentences).tolist()

            base_url = "https://dau-stg.csod.com/ui/lms-learning-details/app/onlineContent/"
            columns = ['KSAB', 
                   'Course 1', 'Course Name 1', 'Score 1', 'Course Link 1',
                   'Course 2', 'Course Name 2', 'Score 2', 'Course Link 2',
                   'Course 3', 'Course Name 3', 'Score 3', 'Course Link 3',
                   'Course 4', 'Course Name 4', 'Score 4', 'Course Link 4',
                   'Course 5', 'Course Name 5', 'Score 5', 'Course Link 5']
            df = pd.DataFrame(columns=columns)


            for query_embedding, query_sentence in zip(query_embeddings, query_sentences):
                res = self.index.query(vector=query_embedding, namespace="course_descriptions", top_k=5, include_values=True)

                # Extract the course IDs, scores, and generate course links
                course_ids = [res_match.id for res_match in res.matches]
                scores = [res_match.score for res_match in res.matches]
                course_names = [course_id_to_name.get(course_id, "Unknown") for course_id in course_ids]
                course_links = [base_url + course_id for course_id in course_ids]

                # Fill in NaN if fewer than 5 matches
                while len(course_ids) < 5:
                    course_ids.append(pd.NA)
                    course_names.append(pd.NA)
                    scores.append(pd.NA)
                    course_links.append(pd.NA)

                # Add data to the DataFrame
                data = []
                for i in range(5):
                    data.extend([course_ids[i], course_names[i], scores[i], course_links[i]])

                df.loc[len(df)] = [query_sentence] + data

            base = os.path.basename(filename)
            name, ext = os.path.splitext(base)
            df.to_excel(f'data\\combined_KSAB_excel\\{name}.xlsx', sheet_name=name, index=False)
            df = df.iloc[0:0]
            print(f"{name} done")

    def embedding_query_with_score(self):
        folderpath = "data\\roles"
        if not os.path.exists("data\\excel"):
            os.mkdir('data\\excel')

        # Load course IDs and descriptions from pickle
        with open('data\\course_data.pkl', 'rb') as f:
            course_id_list, course_descriptions = pickle.load(f)

        # Load course ID to name mapping from pickle
        with open('data\\course_id_to_name.pkl', 'rb') as f:
            course_id_to_name = pickle.load(f)

        self.corpus_sentences = course_descriptions

        # Process each roles file
        for filename in glob.glob(os.path.join(folderpath, '*.xlsx')):
            df_roles = pd.read_excel(filename)
            query_sentences = df_roles['description : String'].tolist()
            query_embeddings = self.model.encode(query_sentences).tolist()

            base_url = "https://dau-stg.csod.com/ui/lms-learning-details/app/onlineContent/"
            columns = ['KSAB', 
                    'Course 1', 'Course Name 1', 'Score 1', 'Course Link 1',
                    'Course 2', 'Course Name 2', 'Score 2', 'Course Link 2',
                    'Course 3', 'Course Name 3', 'Score 3', 'Course Link 3',
                    'Course 4', 'Course Name 4', 'Score 4', 'Course Link 4',
                    'Course 5', 'Course Name 5', 'Score 5', 'Course Link 5']
            df = pd.DataFrame(columns=columns)

            for query_embedding, query_sentence in zip(query_embeddings, query_sentences):
                res = self.index.query(vector=query_embedding, namespace="course_descriptions", top_k=5, include_values=True)

                # Extract the course IDs, scores, and generate course links
                course_ids = [res_match.id for res_match in res.matches]
                scores = [res_match.score for res_match in res.matches]
                course_names = [course_id_to_name.get(course_id, "Unknown") for course_id in course_ids]
                course_links = [base_url + course_id for course_id in course_ids]

                # Fill in NaN if fewer than 5 matches
                while len(course_ids) < 5:
                    course_ids.append(pd.NA)
                    course_names.append(pd.NA)
                    scores.append(pd.NA)
                    course_links.append(pd.NA)

                # Populate DataFrame row
                row_data = [query_sentence]
                for i in range(5):
                    row_data.extend([course_ids[i], course_names[i], scores[i], course_links[i]])

                df.loc[len(df)] = row_data

            base = os.path.basename(filename)
            name, ext = os.path.splitext(base)
            df.to_excel(f'data\\combined_KSAB_excel\\{name}.xlsx', sheet_name=name, index=False)
            df = df.iloc[0:0]
            print(f"{name} done")


        
    def delete_namespace(self, namespace):
        self.index.delete(deleteAll='true', namespace=namespace)


def main():
    pi = PineconeInteraction()
    #pi.corpus_embeddings('course_descriptions.txt')
    #pi.embedding_query_with_score()
    #pi.compare_KSAB()
    #color_identical_cells_in_excel(f'data\\excel')
    #create_pivot_table_from_excel(f'data\\excel')
    pivot_table(f'data\\excel')

if __name__=='__main__':
    main()